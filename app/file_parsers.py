from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable


ALLOWED_EXTENSIONS = {".txt", ".md", ".docx", ".pdf", ".xlsx", ".csv"}


class FileParseError(Exception):
    pass


def _normalize_text(value: str) -> str:
    return value.replace("\x00", "").strip()


def _read_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def _read_csv(path: Path) -> str:
    lines: list[str] = []
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            lines.append("\t".join(item.strip() for item in row))
    return "\n".join(lines)


def _read_docx(path: Path) -> str:
    try:
        from docx import Document  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise FileParseError("python-docx is not installed") from exc

    document = Document(str(path))
    chunks = [paragraph.text for paragraph in document.paragraphs if paragraph.text.strip()]
    return "\n".join(chunks)


def _read_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise FileParseError("pypdf is not installed") from exc

    reader = PdfReader(str(path))
    chunks: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        if text.strip():
            chunks.append(text)
    return "\n".join(chunks)


def _read_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise FileParseError("openpyxl is not installed") from exc

    workbook = load_workbook(str(path), read_only=True, data_only=True)
    chunks: list[str] = []
    for sheet in workbook.worksheets:
        chunks.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = ["" if item is None else str(item) for item in row]
            line = "\t".join(value.strip() for value in values if value is not None)
            if line.strip():
                chunks.append(line)
    return "\n".join(chunks)


def parse_file(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise FileParseError(f"Unsupported format: {suffix}")

    if suffix in {".txt", ".md"}:
        text = _read_text_file(path)
    elif suffix == ".csv":
        text = _read_csv(path)
    elif suffix == ".docx":
        text = _read_docx(path)
    elif suffix == ".pdf":
        text = _read_pdf(path)
    elif suffix == ".xlsx":
        text = _read_xlsx(path)
    else:  # pragma: no cover
        raise FileParseError(f"Unhandled format: {suffix}")

    normalized = _normalize_text(text)
    if not normalized:
        raise FileParseError("No extractable text found in file")
    return normalized


def ensure_allowed_filename(filename: str) -> None:
    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        allowed = ", ".join(sorted(ALLOWED_EXTENSIONS))
        raise FileParseError(f"Extension not allowed. Supported formats: {allowed}")
